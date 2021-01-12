from openpyxl import load_workbook
from pydriller import GitRepository
from pydriller import RepositoryMining

workbook = load_workbook('/usagers3/opmos/Documents/OPM/Mouna Research/BugFixingCommits00.xlsx')
sheets = workbook.sheetnames
index = 0
for sheet in sheets:
    worksheet = workbook.get_sheet_by_name(sheet)
    shaaList = worksheet["B"]
    # Print the contents
    for x in range(2,len(shaaList)):
        print(shaaList[x].value)
        #gr = GitRepository("https://github.com/{}".format(worksheet['A2'].value))
        gr = GitRepository("https://github.com/{}".format(worksheet['A2'].value))
        for commit in RepositoryMining ('https://github.com/{}'.format(worksheet['A2'].value),single ='{}'.format(shaaList[x].value)).traverse_commits():
            #print(commit.hash)
            #print(commit.msg)
            for modified_files in commit.modifications:
                buggy_commits = gr.get_commits_last_modified_lines(commit,modified_files)
                print("       {} ".format(buggy_commits)) # result: (abc, def)

        #for commit in RepositoryMining('../test-repos/test1/',).traverse_commits():
        #print('hash {} authored by {}'.format(commit.hash, commit.author.name))
